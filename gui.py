import PySimpleGUI as sg
import openpyxl
from string import digits
import re

def main():
    workbook=r'C:\Work\GitLocals\development\Parameters.xlsx'
    worksheet="Sheet1"
    wb = openpyxl.load_workbook(workbook)
    ws = wb[worksheet]
    cell_s1="E1"
    cell_s2="E6"

    droplist = gen_list_from_ws(ws=ws,cell_s=cell_s1)
    table    = gen_checkboxs(ws=ws, cell_s=cell_s2)
    print(table)

    
    checkboxs = []


    # ウィンドウのテーマ
    sg.theme('BlueMono')

    # ウィンドウのレイアウト
    table_cols = ["A", "B", "C"]
    cols_width = [100 for _ in range(len(table_cols))]

    table_style = {
        'values': table,
        'headings': table_cols,
        'max_col_width': 100,
        'def_col_width': cols_width,
        'num_rows': 15,
        'auto_size_columns': True,
        'key': '-TABLE-'
    }

    layout = [
            [sg.Text('チェックボックス')],
            [sg.Checkbox("チェックボックス1", key='c1', default=True)],
            [sg.Checkbox("チェックボックス2", key='c2', default=False)],
            [sg.Combo(droplist, default_value="選択して下さい", size=(30,1)) ],
            [sg.Table(**table_style)],


            [sg.Button("Button",key='b1')]
            
        ]

    # ウィンドウオブジェクトの作成
    window = sg.Window('title', layout, size=(300, 300))

    # イベントのループ
    while True:
        # イベントの読み込み
        event, values = window.read()
        print(f'checkbox1 is {values["c1"]}, checkbox2 is {values["c2"]}')


        # ウィンドウの×ボタンクリックで終了
        if event == sg.WIN_CLOSED:
            break

    # ウィンドウ終了処理
    window.close()
    return



#def layout_combo(workbook, worksheet, start_cell="A1", with_title=True):

def gen_checkboxs(ws,cell_s):
    bottom_cell    = return_row_deadend(ws,cell_s)
    most_left_cell = return_col_deadend(ws,cell_s)
    cell_e = return_deadend(bottom_cell=bottom_cell, most_left_cell=most_left_cell)
    table = []
    for row in ws[f"{cell_s}:{cell_e}"]:
        values = []
        for col in row:
            values.append(col.value)
        table.append(values)
    return table

    


def return_deadend(bottom_cell, most_left_cell):
    end_c = re.sub(r"[\d]*$",     "", most_left_cell)
    end_r = re.sub(r"^[A-Za-z]*", "", bottom_cell)
    return end_c + end_r

def gen_list_from_ws(ws,cell_s):
    r,c = cell_s2i(cell_s)
    value = ws.cell(row=r,column=c).value
    if(ws.cell(r,c).value is None):
        raise ValueError("Illegal cell was specified.")

    cell_e = return_row_deadend(ws,cell_s)

    droplist = []
    for row in ws[f"{cell_s}:{cell_e}"]:
        for col in row:
            droplist.append(col.value)

    return droplist


def return_row_deadend(ws,cell_s):
    r,c = cell_s2i(cell_s)
    i=1
    while(True):
        value = ws.cell(row=r+i,column=c).value
        if(value is None):
            max_r = r+i-1
            break
        i+=1
    return cell_i2s(row=max_r, column=c)

def return_col_deadend(ws,cell_s):
    r,c = cell_s2i(cell_s)
    i=1
    while(True):
        value = ws.cell(row=r,column=c+i).value
        if(value is None):
            max_c = c+i-1
            break
        i+=1
    return cell_i2s(row=r, column=max_c)


def cell_s2i(cell:str)->tuple[int,int]:
    table = str.maketrans('', '', digits)
    col =  openpyxl.utils.column_index_from_string(cell.translate(table))
    row = int(re.sub(r"^[A-Za-z]*", "", cell))
    return row,col
    
def cell_i2s(row, column):
    return openpyxl.utils.get_column_letter(column) + str(row)

    
    




if __name__ == "__main__":
    main()