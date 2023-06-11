# １：ライブラリ設定
import fitz # pymupdfライブラリ
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import json
import re
from tqdm import tqdm
import argparse




class Attr:
    def __init__(self,size="*", font="*", color="*", flags="*", indent="*", vertical="*"):
        self.size = size
        self.font = font
        self.color = color
        self.flags = flags
        self.indent = indent
        self.vertical = vertical

    def __eq__(self, other):
        if(other.size != "*" and self.size != other.size):
            return False
        if(other.font != "*" and self.font != other.font):
            return False
        if(other.color != "*" and self.color != other.color):
            return False
        if(other.flags != "*" and self.flags != other.flags):
            return False
        if(other.indent != "*"):
            if(isinstance(other.indent,list) and other.indent[0]<=self.indent<=other.indent[1]):
                pass
            elif(self.indent != other.indent):
                return False
        if(other.vertical != "*"):
            vertical_float = float(other.vertical[1:])
            if(other.vertical[0] == ">"  and self.vertical < vertical_float):
                return False
            if(other.vertical[0] == "<"  and self.vertical > vertical_float):
                return False

        return True
    
#ParegraphクラスがクラスLineを配列のとして格納する
# Lineには、 spanを解析することで分類された属性と文からなる。
class Line():
    def __init__(self, span:object, rule:dict):
        self.attr = Attr(
            size=span.get("size"),
            font=span.get("font"),
            color=span.get("color"),
            flags=span.get("flags"),
            indent=span.get("origin")[0],
            vertical=span.get("origin")[1]
            )
        self.span = span
        self.type = None
        self.text = span.get("text")
        for type, formats in rule.items():
            for format in formats:
                other_attr = Attr(format["size"], format["font"], format["color"], format["flags"], format["indent"], format["vertical"])
                if(self.attr == other_attr):
                    self.type = type
                    break
            else:
                continue
            break
        # Work Around ： MPHYは 章の名前なのか、 図の名前なのかが、フォーマットからは分別できない。
        # 特別措置として、TypeがTitle のとき、 テキストの始まりが Figureの場合は、TypeをFigureNameに変更する
        if(self.type in ["Title4", "Title5"]):
            if(self.text.startswith("Figure") ):
                self.type = "FigureTitles"
            if(self.text.startswith("Table") ):
                self.type = "TableTitles"

    def __eq__(self,other):
        if(self.type == other.type):
            return True
        else:
            return False


#パラグラフは 章の名前と、本文に分類する。本文には 通常文書、リスト、図表名を区別する。
#通常文書、リスト、図表名 が切り替わっタイミングで、切り替わり前のを1つの文字列する
# またその１つの文字列が何なのかを示す属性(ID)を付加する。


def search_keywords(str, keywords):
    result = []
    for i,keyword in enumerate(keywords):
        find_status = str.find(f" {keyword} ")
        if(find_status==-1):
            result.append("-")
        else:
            result.append("x")
    if("x" in result):
        result.append("x")
    else:
        result.append("-")
    return result


class Write:
    def __init__(self, worksheet, file_object, start_row_pos=1, paragraph_col="A", description_col="B",
                 font_name = 'メイリオ',
                 fill_style=PatternFill(patternType='solid', fgColor='ADFF2F')):
        self.ws = worksheet
        self.f  = file_object
        self.row=start_row_pos
        self.p_col = paragraph_col
        self.d_col = description_col
        self.font_name = font_name
        self.fill_style = fill_style
        pass

    def write_title(self, title:list, col, row_inc_val_after_write=1, 
                    apply_style_fill=False, 
                    wrap=False,
                    bold=False, 
                    underline=None, 
                    italic=False):
        cell = self.ws[f"{col}{self.row}"]
        col_num = openpyxl.utils.cell.column_index_from_string(col)
        for i, a_title in enumerate(title):
            cur_col_num = col_num + i
            cur_col_str = openpyxl.utils.cell.get_column_letter(cur_col_num)
            cell = self.ws.cell(column=cur_col_num,row=self.row)
            cell.value  = a_title
            self.ws[cell.coordinate].font = Font(name=self.font_name, bold=bold, underline=underline, italic=italic)
            self.ws.column_dimensions[cur_col_str].width = len(a_title) *1.6
            self.f.write(f'"{a_title}",')
        self.row+=row_inc_val_after_write

    def change_width (self,col,width):
        self.ws.column_dimensions[col].width = width



        
    def write_down(self, value:str, additional_list:list, col, row_inc_val_after_write=1, apply_style_fill=False, wrap=False, bold=False, underline=None, italic=False):
        self.f.write(f',"{str}",')
        self.f.write(",".join(additional_list))
        self.f.write('\n')
        cell = self.ws[f"{col}{self.row}"]
        cell.value = value
        if(apply_style_fill):
            self.ws[cell.coordinate].fill = self.fill_style
        if(wrap):
            cell.alignment = openpyxl.styles.Alignment(wrapText=True)
        self.ws[cell.coordinate].font = Font(name=self.font_name, bold=bold, underline=underline, italic=italic)
        for col_ofs,a_result in enumerate(additional_list):
            cell.offset(column=col_ofs+1,row=0).value=a_result
            if(a_result=="x"):
                self.ws[cell.offset(column=col_ofs+1,row=0).coordinate].fill = self.fill_style
        self.row+=row_inc_val_after_write
        pass

def main():
    
    # 引数の設定
    parser = argparse.ArgumentParser(description='')
    parser.add_argument('filename', help='PDF')
    parser.add_argument('doc_type', help='PDF')
    parser.add_argument('--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'], default='INFO', help='ログレベルを指定します')
    parser.add_argument('--start_page', type=int, default=1, help="開始ページ")
    parser.add_argument('--end_page',   type=int, default=10, help="終了ページ")
    # 引数の解析
    args = parser.parse_args()
    filename=args.filename
    doc_type=args.doc_type
    page_start = args.start_page -1 #88
    page_end   = args.end_page
    #filename = r"C:\Users\phantasm\Downloads\NCB-PCI_Express_Base_5.0r1.0-2019-05-22.pdf"

    # ２：PDFテキストを格納するリスト作成
    txt_list = []

    # ３：PDFファイルを読み込む
    doc = fitz.open(filename)

    # ４：１ページずつテキストデータを取得

    # 番号なしリストの1行目は、　lines の要素が2個 (同リストの事業は別line)
    # block はページに相当する。

    #設定ファイルを読み込む
    config_file = "pdf_w_PyMuPDF_setting.json"
    with open(config_file, encoding='utf-8') as f:
        config_dict = json.load(f)

    structure_dict = config_dict.get(doc_type)
    keywords       = config_dict.get("Table").get("Keywords")


    analyzed_lines = []
    dump_file = "pdf_w_PyMuPDF_dump.json"
    if(page_end > doc.page_count):
        page_end = doc.page_count
    with open (dump_file, "w") as f:
        for page in tqdm(range(page_start, page_end -1)):
            
            # 解析用に json file formatで取得した内容(Page毎)をファイルに書き出す()
                #print(doc[page].get_text("json"))
            try:
                f.write(doc[page].get_text("json"))
            except IndexError:
                break
                
            images = doc[page].get_images()
            if(len(images)):
                try:
                    f.write(doc[page].get_images("json"))
                except TypeError:
                    f.write(f"!!! ERRROR  in page{page}!!!!")


            # PDF情報(Page毎)を Json format で出力した内容を辞書に変換し変数に格納する
            json_load = json.loads(doc[page].get_text("json"))
            pass

            # 現ページの情報をBlockという単位で取り出す
            for block in json_load.get("blocks"):
                #type of the block is "list"
                #print(f'key={key}, value={value}')
                #print(f'block={block}')
                #現Blockの情報を1行ずつ取り出す。
                if(block.get("lines") is None):
                    continue
                for line in block.get("lines"):
                    #print(f'line={line}')
                    # 章の名前と、内容に分離したあと、両方を オブジェクト Paragraph に格納する
                    # ここでBody部分に該当する行はすべて連結する。
                    for span in line.get("spans"):
                        text = span.get("text")
                        size = span.get("size")
                        analyzed_lines.append(Line(span, structure_dict))



    ignore_list = ["Invalid", "FigureBodies", "NoMatch", "flag0", "flag2"]

    debug_file = "pdf_w_PyMuPDF_debug.log"
    with open (debug_file, "w", encoding='utf-8') as f:
        for line in analyzed_lines:
            if(False and line.type == "Invalid"):
                pass
            elif(re.fullmatch(r'\s*', line.text)):
                pass
            else:
                f.write(f'type:{line.type:15s},font:{line.span.get("font"):30s}, size:{line.span.get("size"):17.15f}, color:{line.span.get("color"):8d}, flag:{line.span.get("flags"):2d}, indent:{line.span.get("origin")[0]:19.15f},vertical:{line.span.get("origin")[1]:19.15f}, \t{line.text}\n')

        for line in analyzed_lines:
            if(line.type in ignore_list):
                continue
            f.write(f'type:{line.type:15s},\t{line.text}\n')

    analyzed_lines_wo_invalid = []
    pre_type = None
    for line in analyzed_lines:
        if(line.type in ignore_list):
            pass
        elif(line.type == "Mirror" and pre_type in ignore_list):
            pass
        elif(re.fullmatch(r'\s*',line.text)):
            pass
        else:
            if(line.type =="Mirror"):
                line.type = pre_type
            analyzed_lines_wo_invalid.append(line)
        if(line.type !="Mirror"):
            pre_type = line.type

    result_file = "pdf_w_PyMuPDF_out.csv"
    connect_flag = False
    put_flag     = False
    body_flag    = False

    
    result_xsl = "pdf_w_PyMuPDF_out.xlsx"
    sheet_name = "MySheet"
    paragraph_col   = "A"
    description_col = "B"
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = sheet_name
    ws.freeze_panes = 'A2'
    ws.column_dimensions[description_col].width = 100
    font = Font(name='メイリオ')

    fill = PatternFill(patternType='solid', fgColor='ADFF2F')
    





    with open (result_file, "w", encoding='utf-8') as f:
        wp = Write(worksheet=ws, file_object=f, start_row_pos=1, paragraph_col="A", description_col="B")
        last = len(analyzed_lines_wo_invalid) -1
        str = ""
        pre_type = None

        # Print Title of the table
        table_title = []
        table_title.append("Par. Name")
        table_title.append("Description")
        for keyword in keywords:
            table_title.append(f"{keyword}")
        table_title.append(f"any")
        wp.write_title(table_title, col="A", bold=True)
        wp.change_width(description_col,100)

        f.write(f',\n')
        for i,line in enumerate(analyzed_lines_wo_invalid):
            if(i<last):
                next_line = analyzed_lines_wo_invalid[i+1]
                if(line.type==next_line.type):
                    put_flag = False
                else:
                    if(next_line.type=="BulletPoint"):
                        #リストが連続していても、 ・ が出てくれば一旦出力
                        put_flag = True
                    elif(line.type=="BulletPoint" and next_line.type.startswith("Items")):
                        put_flag = False
                    else:
                        put_flag = True
            else:
                put_flag = True

            if(put_flag):
                str += line.text
                if(line.type == "Bodies"):
                    while(True):
                        m = re.search(r"(?P<match>.*?\.) " , str)
                        if(m):
                            str = re.sub(r"^.*?\. ", "", str)
                            match = m.group("match")
                            # 両端にある " " を削除
                            match = match.strip(" ")
                            search_result = search_keywords(match,keywords)
                            wp.write_down(value=match, additional_list=search_result, col=description_col, wrap=True)

                        else:
                            # 両端にある " " を削除
                            str = str.strip(" ")
                            if(str!=""):
                                search_result = search_keywords(str,keywords)
                                wp.write_down(value=str, additional_list=search_result, col=description_col, wrap=True)
                            break
                elif(line.type.startswith("Items")):
                    search_result = search_keywords(str,keywords)
                    put_indent = ""
                    if(line.type=="Items2"):
                        put_indent = "    "
                    wp.write_down(value=put_indent+str, additional_list=search_result, col=description_col, wrap=True)
                elif(line.type.startswith("Title")):
                    wp.write_down(value=str, additional_list=[], col=paragraph_col, wrap=False)
                elif(str!=""):
                    search_result = search_keywords(str,keywords)
                    wp.write_down(value=str, additional_list=search_result, apply_style_fill=True, col=description_col, wrap=False)
                str = ""
            else:
                str += line.text

    
    wb.save(result_xsl)
                    

                



    pass



if __name__ == "__main__":
    main()
