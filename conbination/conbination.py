import openpyxl
import re
import numpy as np
import itertools
import pandas as pd


def expand_to_list(str_or_other):
    """ 文字列をルールに従って展開(可能なら)し、リストで返す

    Args:
        str_or_other (_type_): _description_

    Returns:
        _type_: _description_
    """
    
    if(str_or_other=="None"):
        return [""]
    # 100 to 200 by 50 の形式であれば展開する。前提としてそれぞれが整数
    m = re.search(r"([\d]+) to ([\d]+) by ([\d]+)", str_or_other)
    if(m):
        start  = int(m.group(1))
        end    = int(m.group(2))
        delta  = int(m.group(3))
        return list(range(start, end+delta, delta))
    else:
        return re.split(r"\s*,\s*", str_or_other)

def df_swapped_to_excel(df):
    """ df の行と列を入れ替え、エクセルに保存

    Args:
        df (_type_): データフレーム
    """
    matrix = df.to_numpy()
    # 行列を入れ替える
    transposed_matrix = np.transpose(matrix)
    # Excelファイルを作成して行列を書き込む
    wb = openpyxl.Workbook()
    ws = wb.active
    # 行列をExcelに書き込む
    for i, row in enumerate(transposed_matrix):
        row_list = row.tolist()
        row_list.insert(0,list(df.columns)[i])
        ws.append(row_list)
    # Excelファイルを保存
    wb.save('excel_file.xlsx')

    
PRESET_DICT = {
    "P0" : (0.10, 0.90),
    "P1" : (0.15, 0.85),
    "P2" : (0.20, 0.80),
    "P3" : (0.25, 0.75),
    "P4" : (0.30, 0.70),
    "P5" : (0.35, 0.65),
    "P6" : (0.40, 0.60),
    "P7" : (0.45, 0.55),
    "P8" : (0.50, 0.50),
    "P9" : (0.55, 0.45),
    
}
def main():
    workbook="conditions.xlsx"
    worksheet="Sheet1"
    wb = openpyxl.load_workbook(workbook)
    ws = wb[worksheet]

    expanded_table_list = []
    #ワークシートを1列ずつ読み込み
    columns = []
    for i_col, cols in enumerate(ws.iter_cols()):
        expanded_table = []
        for i_row, cell in enumerate(cols):
            if(i_row==0):
                continue
            if(i_col==0):
                columns.append(cell.value)
                continue
            else:
                #行頭、行末の空白削除、連続する空白を1つに置換
                value = re.sub(r'\s+', ' ', str(cell.value).lstrip().rstrip() )
                # 値を展開した結果(配列)を別の配列に格納
                values = expand_to_list(value)
                expanded_table.append(values)
        if(i_col!=0):
            expanded_table_list.append(expanded_table)
        
    df_list = []
    test_no = 1
    POS_PRESET = 3
    POS_TAP1   = 4
    POS_TAP3   = 5
    for expanded_table in expanded_table_list:
        # 1列分のデータから全組み合わせを生成し、組み合わせを1つずつデータフレームを格納
        for params in list(itertools.product(*expanded_table)):
            params = list(params) #tuple をリストに変更
            params[0] = test_no # No は新規に振り直し
            params[POS_TAP1], params[POS_TAP3] = PRESET_DICT[params[POS_PRESET]]
            #テスト条件1つを１つのデータフレームとして保存
            df_sub = pd.DataFrame([params], columns=columns)
            #そのデータフレームをリストに格納
            df_list.append(df_sub)
            test_no += 1

    #リストに格納したデータフレームを結合
    df = pd.concat(df_list)
    df_swapped_to_excel(df)

    #生成したテストの本数や所要時間(概算)を表示
    total_test_number = len(df)
    run_time_per_a_test = 4
    tat_hour = (run_time_per_a_test * total_test_number)//60
    tat_min  = (run_time_per_a_test * total_test_number)%60
    print(f'生成された組合せ数は {total_test_number} でした。')
    print(f'{tat_hour}時間{tat_min}分かかる見込みです。({run_time_per_a_test}分/本で計算)')
    pass
    



if __name__ == "__main__":
    main()



