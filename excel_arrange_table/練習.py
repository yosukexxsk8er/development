from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
import math

import numpy as np
import os

from  excel_arrange_table.num_to_col_name import num_to_col_name

def arrange_img(img_path_table, worksheet):
    """ table[n] の右方向に並べ、 table[n+1]は次の行に並べる
    画像の大きさに合わせてセルの大きさを変更し、
    1つのセルに1つの画像が収まるよう
    
    Args:
        img_path_table (_type_): 画像へのパスを格納した2次元配列
        worksheet (_type_): ワークシート
    """
    ws = worksheet
    #画像の間隔を設定
    col_padding_num = 1 #とある画像(の左上座標)と右隣の画像(の左上座標)の間隔
    row_padding_num = 2 #とある画像(の左上座標)と下の画像(の左上座標)の間隔。画像分と画像リンク分で2行
    #画像の大きさから適切なセルの大きさを算出ための係数の導入式
    margin = 1.01
    width_ratio_fom_img_to_cell  = ( 57 / 400) * margin # 過去の調整の結果から比率を算出
    height_ratio_fom_img_to_cell = (170 / 200) * margin # 過去の調整の結果から比率を算出
    #貼り付ける画像(全画像の大きさは同じ前提) の大きさを取得し、上記の比率をかけて、セルの大きさを決定
    cell_width  = math.ceil(Image(img_path_table[0][0]).width  *  width_ratio_fom_img_to_cell)  #400 -> 57
    cell_height = math.ceil(Image(img_path_table[0][0]).height *  height_ratio_fom_img_to_cell) #200 -> 170
    # 画像パスを記載するセルの書式を設定
    title_format = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    title_font = Font(name="Meiryo", bold=True)

    i_col = 1
    i_row = 1
    for img_path_list in img_path_table:
        for img_path in img_path_list:
            # 画像を読み込む
            img = Image(img_path)
            cell_name = f"{num_to_col_name(i_col)}{i_row}"
            #まずは画像の名称(ハイパーリング)を記載
            ws[cell_name] = f'=HYPERLINK("{img_path}","{os.path.basename(img_path)}")'
            ws[cell_name].fill = title_format
            ws[cell_name].font = title_font 
            #画像を貼り付け
            cell_name = f"{num_to_col_name(i_col)}{i_row+1}" #画像挿入のためターゲットのセルを1つ下の行へ移動
            ws.column_dimensions[f'{num_to_col_name(i_col)}'].width = cell_width
            ws.row_dimensions[i_row+1].height = cell_height
            ws.add_image(img, cell_name)
            #次のループので貼り付ける列へ移動
            i_col += col_padding_num
            pass
        #次のループ(次の行)のターゲット位置(次の行、列ははじめに戻す)
        i_row += row_padding_num
        i_col = 1


def main():
    wb = Workbook()
    ws = wb.active

    pattern_list = ["CLK", "PRBS9", "MCP-IR"]
    Preset_list = list(np.char.add("P", np.arange(10).astype(str)))
    dir = "outputs"
    
    # 2次元配列で画像ファイルパスの一覧を作成
    img_path_table = []
    for preset in Preset_list:
        img_path_list = []
        for pattern in pattern_list:
            img_path_list.append(f'{dir}/{pattern}_{preset}.png')
        img_path_table.append(img_path_list)
            

    arrange_img(img_path_table=img_path_table, worksheet=ws)
    pass

    pass
    output_file = "output.xlsx"
    wb.save(output_file)



if __name__ == "__main__":
    main()
